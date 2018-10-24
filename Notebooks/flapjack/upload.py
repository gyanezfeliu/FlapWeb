import openpyxl as opxl
import numpy as np
import flapjack
from flapjack import tables as tbc
import os
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import pandas.io.sql as psql
#import Tkinter, tkFileDialog

# ---
# Format specific functions to load measurement data from sheet of input Excel file
# ---
def find_in_sublists(lst, value):
    for sub_i, sublist in enumerate(lst):
        try:
            return (sub_i, sublist.index(value))
        except ValueError:
            pass
    raise ValueError('%s is not in lists' % value)


def load_bmg_data(sheet, sample_ids, session, engine):
    rownames = ['CFP', 'YFP', 'OD']
    nsets = len(rownames)
    rawdata = [[cell.value for cell in row] for row in sheet.iter_rows()] #list(csv.reader(open(filename, 'rU')))

    # Work out location of header row, and column/row numbers
    try:
        (headerrow,ccol) = find_in_sublists(rawdata,'Well Col')
        (headerrow,crow) = find_in_sublists(rawdata,'Well Row')
    except ValueError:
        (headerrow,ccol) = find_in_sublists(rawdata,'Well\nCol')
        (headerrow,crow) = find_in_sublists(rawdata,'Well\nRow')

    # Time is on row below headers
    timerow = headerrow+1
    # Data is on rows below time
    datarow = timerow+1

    # Find where the data is in each row by looking for 1st 'Raw Data'
    headers = rawdata[headerrow]
    cdata = next((i for i,v in enumerate(headers) if 'Raw Data' in v))
    r = rawdata[datarow:]
    r1 = r[1]

    # Pull out row,col and data for each row
    # Try combinations of alpha/numeric for row/col
    try:
        #rows = [(ord(row[crow].upper())-64, int(row[ccol]), map(float, row[cdata:])) for row in rawdata[datarow:]]
        rows = [(ord(row[crow].upper())-64, int(row[ccol]), [float(v) for v in row[cdata:]]) for row in rawdata[datarow:]]
    except ValueError:
        #rows = [(int(row[ccol]), ord(row[crow].upper())-64, map(float, row[cdata:])) for row in rawdata[datarow:]]
        rows = [(int(row[ccol]), ord(row[crow].upper())-64, [float(v) for v in row[cdata:]]) for row in rawdata[datarow:]]

    # Pull out time of each data point
    times = rawdata[timerow]
    times = times[cdata:]

    assert len(list(rows[0][2])) % nsets == 0
    nsteps = len(list(rows[0][2])) // nsets


    data = [['','row','col','t'] + rownames]
    

    n = 0 
    for row, col, vals in rows:
        sidx = (row-1)*12 + col-1
        for i in range(nsteps):
            t = float(times[i])
            line = [n, row, col, t]
            for j in range(nsets):
                k = j*nsteps + i
                line.append(vals[k])
                m = tbc.measurement(name = rownames[j], value = vals[k], time = t, sample_id = int(sample_ids[sidx]))
                session.add(m)
            data.append(line)
        n = n+1
    session.commit()

def load_synergy_data(sheet, samples_id, session, medidas):
    print('Loading measurements...')
    # Map measurement names to standard names
    name_map = {'OD600:600':'OD', 'RFP-YFP:500/27,540/25':'YFP', 'CFP:420/50,485/20':'CFP',
    'RFP-YFP:585/10,620/15':'RFP'}

    # list of tuples ordered as measurement name, row where it is, col
    # used to get in wich rows are the begining for each table of a specific measurement
    lista_rows = [(celda.value, celda.row, opxl.utils.column_index_from_string(celda.column)) for celda in sheet['A'] if celda.value in medidas]
    d_tablas = {}
    # uses the get values to collect the data from the excel to a np array
    # d_tablas keys are measurement names cointaining a tuple with a value and a time
    for dato in lista_rows:
        data, time = get_synergy_values(sheet, dato[1] + 3, dato[2] + 3 , dato[2] + 1)
        d_tablas[dato[0]] = (data, time)

    # for each measurement, loads for each time the value of every sample on that moment
    # then goes for the next time, and so on. Also relates to the specific sample, commits to the db
    for name in d_tablas.keys():
        for time_index in range(len(d_tablas[name][1])):
            i = 1
            for value in d_tablas[name][0][time_index]:
                m = tbc.measurement(name = name_map[name], \
                                    value = float(value), time = float(d_tablas[name][1][time_index]),
                sample_id = int(samples_id[i-1]))
                i+=1
                session.add(m)
    session.commit()

def get_synergy_values(ws,row_start,column_start, time_column):
    """
    gets a sheet, a row and column (ints) as base pointers, the time column (int)
    check how big the table is, and starts to get for each time the measured value for each column (a different sample)
    then returns the data and time numpy arrays
    """

    # Get limits of data region in sheet
    maxrow,maxcol = row_start, column_start
    while ws.cell(row=maxrow, column=maxcol).value != None:
        maxrow+=1
    maxrow -= 1
    while ws.cell(row=maxrow , column=maxcol).value != None:
        maxcol+=1
    maxcol -= 1

    data = np.zeros((maxrow-row_start+1,maxcol-column_start+1))
    time = np.zeros((maxrow-row_start+1))
    for i in range(maxrow - row_start + 1):
        time_cell = ws.cell(row=i+row_start, column=time_column)
        time_cell.number_format = 'Number'
        time_aux = time_cell.value
            #transform time values in hour fractions units
        if type(time_aux)==float:
            time[i] = time_aux*24.
        else:
            time[i] = time_aux.hour + time_aux.minute/60. + time_aux.second/3600.
            # Handle bad time formatting with rollover after 24 hours
            if i>0 and time[i]<=time[i-1]:
                time[i] += time[i-1]
        
        for j in range(maxcol - column_start + 1):
            data[i,j] = ws.cell(row=i+row_start, column=j+column_start).value

    return(data, time)

# 
# Functions to extract metadata from 12x8 tables in sheets of Excel file
#
def small_table(sht, row):
    # gets a sheet and a row as top reference of a table to be read
    # read the info intothe table and returns a dictionary with the values (any type) and with a 'name' key tat describes the table
    dic = {}
    c  = 1
    dic['name'] = sht[row][0].value
    for cell_row in range(row + 1, row + 9):
        for cell_col in range (1, 13):
            if sht[cell_row][cell_col].value == None:
                dic[c] = 0
            dic[c] = sht[cell_row][cell_col].value
            c += 1
    return dic

def table_values(ws, row):
    # Get a list of values in a table (12x8) at given row position
    name = ws[row][0].value
    vals = []
    for cell_row in range(row + 1, row + 9):
        for cell_col in range (1, 13):
            val = ws[cell_row][cell_col].value
            if not val:
                vals.append(0)
            else:
                vals.append(val)
    return name,vals

def get_all_tables(ws):
    # Return a list of tables (12x8) as lists in worksheet
    # Tables are assumed to be ordered vertically
    length = len(ws['A'])
    ntables = (length + 1) // 10
    table_list = []
    names_list = []
    for i in range(ntables):
        name,values = table_values(ws, i*10+1)
        table_list.append(values)
        names_list.append(name)
    return names_list,table_list 

def sheet_reader(sht):
    # reads many small tables and returns a list with all the dictionaries
    # being the keys the number of the sample on the plate 1-96 and the value the value of the cell
    if not sht:
        return []
    else:
        largo = len(sht['A'])
        n_tablas = (largo + 1) // 10
        tables_list = []
        for nro_tabla in range(n_tablas):
            tables_list.append(small_table(sht, nro_tabla * 10 + 1))
        return tables_list 

def load_meta_info(wb, experiment_name, machine_name, session, engine):
    ''' 
    Loads the data of the selected file and commits the changes on the database
    '''
    print('Loading metadata...')

    # loads and commits the new experiment
    exp = tbc.experiment(name=experiment_name, machine=machine_name)
    session.add(exp)
    session.commit()
    # get the exp id to make the relations between samples to this specific experiment id
    exp_id = psql.read_sql_table('experiment', engine).id.values[-1]

    # Get inducer info
    _,dna_tables = get_all_tables(wb['DNA'])
    if 'Inducers' in wb.sheetnames:
        inducer_names,inducer_tables = get_all_tables(wb['Inducers'])
    else:
        inducer_tables = []
        inducer_names = []

    #creates the 96 samples, with the specified media and strain
    _,media_table = table_values(wb['Media'], 1)
    _,strain_table = table_values(wb['Strains'], 1)
    for i in range(96):
        sample = tbc.sample(col=i%12+1, row=i//12+1, \
                                experiment_id = int(exp_id), \
                                media = media_table[i], \
                                strain = strain_table[i])
        # Add inducer concentrations to sample
        for i in range(len(inducer_names)):
            inducer_concs = inducer_tables[i]
            inducer_name = inducer_names[i]
            setattr(sample, inducer_name, float(inducer_concs[j]))
        session.add(sample)
    session.commit()

    # Get the ids of the samples just committed (is this safe?)
    sample_ids = psql.read_sql_table('sample', engine).id.values[-96:]

    # Check for existing dna
    existing_dna = psql.read_sql_table('dna', engine)
    # Link dna to samples via vector table, creating new dnas if necessary
    for i in range(len(dna_tables)):
        for j in range(96):
            dna_name = str(dna_tables[i][j])
            if dna_name!='None':
                if dna_name not in existing_dna.name.values:
                    # Create dna
                    session.add(tbc.dna(name=dna_name, sequence=''))
                    session.commit()
                    existing_dna = psql.read_sql_table('dna', engine)
                # Create vector linking plasmid to sample
                dna_id = existing_dna[existing_dna.name==dna_name].id.values[0]
                vector = tbc.vector(dna_id=int(dna_id), sample_id=int(sample_ids[j]))
                session.add(vector)
                session.commit()
    session.commit()

    # Return list of samples loaded
    return sample_ids

#
# Load data and metadata using the above functions
#
def load(file_name, file_format, session, engine, medidas, experiment_name=None, machine_name=None):
    print('--------------------------------------------------------------------')
    wb = opxl.load_workbook(filename = file_name, data_only=True)
    existing_expt_names = psql.read_sql_table('experiment', engine).name.values
    if not experiment_name:
        experiment_name = os.path.basename(file_name) #.split('/')[-1].split('.')[0]
    # Check if experiment is already in db
    if experiment_name in existing_expt_names:
        print('Experiment ', experiment_name, ' already exists in database. Have you already uploaded this file, or one with the same name? Skipping file.')
    else:
        print('Uploading data from file '+file_name)
        st = wb['Data']

        if 'synergy' in file_format:
            if not machine_name:
                machine_name = st['B'][8].value + str(st['B'][9].value)
            samples_id = load_meta_info(wb, experiment_name, machine_name, session, engine)
            load_synergy_data(st, samples_id, session, medidas)
        elif 'bmg' in file_format:
            if not machine_name:
                machine_name = 'bmg'
            samples_id = load_meta_info(wb, experiment_name, machine_name, session, engine)
            load_bmg_data(st, samples_id, session, engine)



































        else:
            print('File format %s not supported'%file_format)


    print('--------------------------------------------------------------------')

