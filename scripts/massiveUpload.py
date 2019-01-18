import pandas as pd
import numpy as np

import json
import os
import openpyxl as opxl
from itertools import islice
from LoadData.models import Experiment, Sample, Dna, Vector, Measurement, Inducer, LoadProcess

def lista_rows(ws, medidas):
    lista_rows = [(celda.value, celda.row, opxl.utils.column_index_from_string(celda.column))
                  for celda in ws['A']
                  if celda.value in medidas]
    return lista_rows

# Changes time's format from datetime to fraction
def fix_synergy_time(df):
    t = np.array([])
    for i, value in enumerate(df['Time']):
        if i > 0:
            if df['Time'].iloc[i].hour < df['Time'].iloc[i-1].hour:
                t = np.append(t, [24 + value.hour + value.minute/60 + value.second/3600])
            else:
                t = np.append(t, [value.hour + value.minute/60 + value.second/3600])
        else:
            t = np.append(t, [value.hour + value.minute/60 + value.second/3600])
    df['Time'] = t

# Cleans the main df
def clean_synergy_data(names, df, lista_rows):
    df_cons = pd.DataFrame(columns=df.columns)
    for i, value in enumerate(lista_rows):
        if i == 0:
            df2 = pd.DataFrame(df.iloc[0:lista_rows[i][1] - 3])
        else:
            df2 = pd.DataFrame(df.iloc[lista_rows[i-1][1] + 1:lista_rows[i][1] - 3])
        df2['name'] = names[i]
        fix_synergy_time(df2)
        df_cons = df_cons.append(df2, ignore_index=True, sort=False)
    return df_cons

def find_in_sublists(lst, value):
    for sub_i, sublist in enumerate(lst):
        try:
            return (sub_i, sublist.index(value))
        except ValueError:
            pass
    raise ValueError('%s is not in lists' % value)

def get_all_tables(ws):
    # Return a list of tables (8x12) as lists in worksheet
    # Tables are assumed to be ordered vertically
    length = len(ws['A'])
    ntables = (length + 1) // 10
    table_list = []
    names_list = []
    for i in range(ntables):
        name,values = table_values(ws, i*10+1)
        table_list.append(values)
        names_list.append(name)
    return names_list,table_list

def table_values(ws, row):
    name = ws[row][0].value
    vals = []
    for cell_row in range(row + 1, row + 9):
        for cell_col in range (1, 13):
            val = ws[cell_row][cell_col].value
            if not val:
                vals.append(0)
            else:
                vals.append(val)
    df_vals = pd.DataFrame(np.reshape(np.array(vals), (8,12)),
                               index=['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'],
                               columns=range(1,13))
    return name, df_vals

def clean_bmg_data(times, rows_data, names, columns):
    l = int(len(rows_data[0])/3)

    t_CFP = times[:l]
    CFP = [r[:l] for r in rows_data]
    t_YFP = times[l:160]
    YFP = [r[l:160] for r in rows_data]
    t_OD = times[160:]
    OD = [r[160:] for r in rows_data]

    list_data = [CFP, YFP, OD]
    list_time = [t_CFP, t_YFP, t_OD]

    # List with time series for 3 measurements
    cols = columns[:]
    cols.insert(0, 'Time')
    cols.append('name')

    df_cons = pd.DataFrame(columns= cols)
    for i, value in enumerate(list_data):
        df = pd.DataFrame(value)
        df = df.transpose()
        df.columns = columns
        df.insert(0, 'Time', list_time[i])
        df['name'] = names[i]
        df_cons = df_cons.append(df, ignore_index=True)
    return df_cons

def load_bmg_data(ws, columns):
    rawdata = [[cell.value for cell in row] for row in ws.iter_rows()]

    # Work out location of header row, and column/row numbers
    try:
        (headerrow,ccol) = find_in_sublists(rawdata,'Well Col')
        (headerrow,crow) = find_in_sublists(rawdata,'Well Row')
    except ValueError:
        (headerrow,ccol) = find_in_sublists(rawdata,'Well\nCol')
        (headerrow,crow) = find_in_sublists(rawdata,'Well\nRow')

    # Time is on row below headers
    timerow = headerrow+1
    # Data is on rows below time
    datarow = timerow+1
    headers = rawdata[headerrow]

    # Find where the data is in each row by looking for 1st 'Raw Data'
    cdata = next((i for i,v in enumerate(headers) if 'Raw Data' in v))

    # Pull out row,col and data for each row
    # Try combinations of alpha/numeric for row/col
    try:
        rows = [(ord(row[crow].upper())-64, int(row[ccol]), [float(v) for v in row[cdata:]]) for row in rawdata[datarow:]]
    except ValueError:
        rows = [(int(row[ccol]), ord(row[crow].upper())-64, [float(v) for v in row[cdata:]]) for row in rawdata[datarow:]]

    # Pull out time of each data point
    times = rawdata[timerow][cdata:]
    # Put just data apart
    rows_data = [r[2] for r in rows]
    names = ['CFP', 'YFP', 'OD']

    df_cons = clean_bmg_data(times, rows_data, names, columns)

    return df_cons

def load_synergy_data(medidas, ws):
    rows_ini = lista_rows(ws, medidas)
    ws.delete_rows(0, rows_ini[0][1] + 1)
    rows = lista_rows(ws, medidas)

    data = ws.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, columns=cols)
    df = df.drop('T° OD600:600', axis=1)

    name_map = {'OD600:600':'OD', 'RFP-YFP:500/27,540/25':'YFP', 'CFP:420/50,485/20':'CFP', 'RFP-YFP:585/10,620/15':'RFP'}
    names = [name_map[rows_ini[i][0]] for i in range(len(rows_ini)-1)]

    df_cons = clean_synergy_data(names, df, rows)
    return df_cons

def load_meta_info(wb, experiment_name, machine_name):
    e = Experiment(name=experiment_name, machine=machine_name)
    e.save()

    # Dictionary where dataframes will be stored and send
    meta_dict = {'Strain_name':'', 'Strain_df':'',
                 'Media_name':'', 'Media_df':'',
                 'DNA_name':'', 'DNA_df':'',
                 'Inducer_name':'', 'Inducer_df':''}

    meta_dict['Strain_name'], meta_dict['Strain_df'] = table_values(wb['Strains'], 1)
    meta_dict['Media_name'], meta_dict['Media_df'] = table_values(wb['Media'], 1)
    meta_dict['DNA_name'], meta_dict['DNA_df'] =  get_all_tables(wb['DNA'])

    if 'Inducers' in wb.sheetnames:
        meta_dict['Inducer_name'], meta_dict['Inducer_df'] = get_all_tables(wb['Inducers'])
    else:
        meta_dict['Inducer_name'], meta_dict['Inducer_df'] = [], []

    return meta_dict, e

def upload_data(e, dict_meta, df_cons, columns):
    for index, value in enumerate(columns):
        row = value[0]
        col = int(value[1:])
        existing_dna = [i.name for i in Dna.objects.all()]

        # Metadata value for each well
        s_strain = dict_meta['Strain_df'].at[row, col]
        s_media = dict_meta['Media_df'].at[row, col]
        s = Sample(experiment=e, row=row, col=col, media=s_media, strain=s_strain)
        s.save()

        for df_dna in dict_meta['DNA_df']:
            s_dna = df_dna.at[row, col]
            if s_dna != 'None':
                if s_dna not in existing_dna:
                    d = Dna(name=s_dna, sboluri='')
                    d.save()
                    v = Vector(dna=d, sample=s)
                    v.save()
                    pass
                else:
                    d = Dna.objects.filter(name__exact=s_dna)[0]
                    v = Vector(dna=d, sample=s)
                    v.save()
                    pass

        if len(dict_meta['Inducer_df']) > 0:
            for i, df_ind in enumerate(dict_meta['Inducer_df']):
                conc = df_ind.at[row, col]
                ind_name = dict_meta['Inducer_name'][i]
                i = Inducer(sample=s, concentration=conc, pubchemid=ind_name)
                i.save()

        # Data value for each well
        for i, val in enumerate(df_cons[row+str(col)]):
            m_name = df_cons['name'].iloc[i]
            m_value = val
            m_time = df_cons['Time'].iloc[i]
            m = Measurement(sample=s, name=m_name, value=m_value, time=m_time)
            m.save()

def load_from_file(route, file_format, columns, medidas):
    files = os.listdir(route)
    for file_name in files:
        if '.xlsx' in file_name:
            wb = opxl.load_workbook(filename = route + file_name, data_only=True)
            ws = wb['Data']
            existing_exp = [i.name for i in Experiment.objects.all()]

            experiment_name = os.path.basename(file_name).split('/')[-1].split('.')[0]
            if experiment_name not in existing_exp:
                if file_format == 'bmg':
                    machine_name = 'bmg'
                    df_cons = load_bmg_data(ws, columns)
                elif file_format == 'synergy':
                    machine_name = ws['B'][8].value + str(ws['B'][9].value)
                    df_cons = load_synergy_data(medidas, ws)

                dict_meta, e = load_meta_info(wb, experiment_name, machine_name)
                upload_data(e, dict_meta, df_cons, columns)
            else:
                # Throws already existent experiment message
                pass
        else:
            # Throws an error message
            pass

columns = [x+str(y) for x in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'] for y in range(1,13)]
medidas = ['OD600:600', 'RFP-YFP:585/10,620/15', 'RFP-YFP:500/27,540/25', 'CFP:420/50,485/20', 'Results']

load_from_file('uploads/datafiles/bmg/to_upload/', 'bmg', columns, medidas)
load_from_file('uploads/datafiles/synergy/to_upload/', 'synergy', columns, medidas)
