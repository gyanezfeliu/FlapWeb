## OLD VERSION, USING DATAFRAMES

# Models
from LoadData.models import Experiment, Sample, Dna, Vector, Measurement, Inducer, LoadProcess
# Data handling
import pandas as pd
import numpy as np
import json
import os
import openpyxl as opxl
from itertools import islice
import csv
import subprocess

def fix_synergy_time(df):
    t = np.array([])
    for i, value in enumerate(df['Time']):
        if i > 0:
            if df['Time'].iloc[i].hour < df['Time'].iloc[i-1].hour:
                t = np.append(t, [24 + value.hour + value.minute/60 + value.second/3600])
            else:
                t = np.append(t, [value.hour + value.minute/60 + value.second/3600])
        else:
            t = np.append(t, [value.hour + value.minute/60 + value.second/3600])
    df['Time'] = t

def run(*args):

    id_data = int(args[0])
    metadata = LoadProcess.objects.get(id=id_data).content
    file_route = LoadProcess.objects.get(id=id_data).file

    # Cargo Metadata
    data = json.loads(metadata)
    df_json = pd.DataFrame(data)
    columns = [x+str(y) for x in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'] for y in range(1,13)]
    df_json.columns = columns
    df_json.index = ['Strain', 'Media', 'DNA']

    experiment_name = os.path.basename(file_route).split('/')[-1].split('.')[0]
    #medidas = ['OD600:600', 'RFP-YFP:585/10,620/15', 'RFP-YFP:500/27,540/25', 'CFP:420/50,485/20', 'Results']
    # Para Isaac
    medidas = ['OD600:600', 'YFP:500/27,540/25', 'CFP:420/50,485/20', 'Results']

    wb = opxl.load_workbook(filename = file_route, data_only=True)
    ws = wb['Data']
    machine_name = ws['B'][8].value + str(ws['B'][9].value)

    name_map = {'OD600:600':'OD', 'YFP:500/27,540/25':'YFP', 'CFP:420/50,485/20':'CFP'}

    # Cargo data
    lista_rows = [(celda.value, celda.row, opxl.utils.column_index_from_string(celda.column))
                for celda in ws['A']
                if celda.value in medidas]

    ws.delete_rows(0, lista_rows[0][1] + 1)

    lista_rows2 = [(celda.value, celda.row, opxl.utils.column_index_from_string(celda.column))
                for celda in ws['A']
                if celda.value in medidas]

    data = ws.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, columns=cols)

    df = df.drop('T° OD600:600', axis=1)

    df_OD = pd.DataFrame(df.iloc[0:lista_rows2[0][1] - 3])
    df_OD['name'] = 'OD'
    fix_synergy_time(df_OD)
    df_OD.index = range(97)

    df_RFP = pd.DataFrame(df.iloc[lista_rows2[0][1] + 1:lista_rows2[1][1] - 3])
    df_RFP['name'] = 'RFP'
    fix_synergy_time(df_RFP)
    df_RFP.index = range(97)

    df_YFP = pd.DataFrame(df.iloc[lista_rows2[1][1] + 1:lista_rows2[2][1] - 3])
    df_YFP['name'] = 'YFP'
    fix_synergy_time(df_YFP)
    df_YFP.index = range(97)

    df_CFP = pd.DataFrame(df.iloc[lista_rows2[2][1] + 1:lista_rows2[3][1] - 3])
    df_CFP['name'] = 'CFP'
    fix_synergy_time(df_CFP)
    df_CFP.index = range(97)

    dfs = [df_OD, df_RFP, df_YFP, df_CFP]

    # 1) Experiment
    e = Experiment(name=experiment_name, machine=machine_name)
    e.save()

    # 2) Sample
    # get experiment_id
    # row, col, media, strain

    # 3) DNA
    # name

    # Empiezo a recorrer los platillos que están en df_json: indexes: strain, media, dna:
    for col_name, col_serie in df_json.iteritems():
        # experiment_id
        existing_dna = [i.name for i in Dna.objects.all()]

        plate_row = col_name[0]
        plate_col = col_name[1:]
        st = col_serie['Strain']['value']
        med = col_serie['Media']['value']

        s = Sample(experiment=e, row=plate_row, col=plate_col, media=med, strain=st)
        s.save()

        DNA_name = col_serie['DNA']['value']
        if DNA_name != 'None':
            if DNA_name not in existing_dna:
                d = Dna(name=DNA_name, sboluri='')
                d.save()
                v = Vector(dna=d, sample=s)
                v.save()
            else:
                d = Dna.objects.filter(name__exact=DNA_name)[0]
                v = Vector(dna=d, sample=s)
                v.save()
        # 4) Measurement
        # name, value, time
        for df in dfs:
            for i, value in enumerate(df[col_name]):
                nam = df['name'].iloc[i]
                val = value
                t = df['Time'].iloc[i]
                m = Measurement(sample=s, name=nam, value=val, time=t)
                m.save()
